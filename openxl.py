

import openpyxl
from pprint import pprint as pp

wb = openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\TC Cert Report.xlsx")
#print(wb.sheetnames)
sheet = wb['YR-2020']

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        while col < sheet.max_column:
            print(sheet.cell(row, col).value, end=' ')
            break
        continue
    print()

