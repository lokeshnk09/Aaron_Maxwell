import openpyxl
from pprint import pprint as pp

wb = openpyxl.load_workbook("/home/lokesh/Documents/Cable.xlsx")
#print(wb.sheetnames)
sheet = wb['Sheet1']

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        while col < sheet.max_column:
            print(sheet.cell(row, col).value, end=' ')
            break
        continue
    print()











        # state = sheet['A' + str(rowobj)].value




# for row in sheet.values:
#     for value in row:
#         print(value)
#     print()



# row = list(sheet.rows)[2]
#
#
# # for rowobjects in tuple(sheet.rows):
# #     for cellval in rowobjects:
# #         print(cellval.coordinate, cellval.value)
# #     print()
#
# #
# # c = range(0, sheet.max_column + 1)
# # r = range(1, sheet.max_row + 1)
# #
# #
# print(help(openpyxl))
#
# #
# # #
# # for c in range(1, sheet.max_column + 1):
# #
#
#
# # for rowobj in sheet.columns:
# #     for cellobj in rowobj:
# #         print(cellobj.coordinate, cellobj.value)
# #     print()
#
# # for item in col[2]:
# #     n_col = sheet.columns
# #     print(item.value)
#
#
# # t = tuple(sheet['A1':'A334'])
# # for rowofcellobj in list(sheet['A1']]):
# #     for cellobj in rowofcellobj:
# #         print(cellobj.value)
# #     # for cellobj in list(sheet.columns)[1]:
# #     #     print(cellobj.value)
# #
#
